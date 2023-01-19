import psycopg2
import pandas as pd
import sys
from sqlalchemy import create_engine
import pandas as pd
import os
import unidecode
import re
from datetime import date, timedelta, datetime
from utils import (
    DATABASE_INFO,
    findfiles,
    no_accent_vietnamese,
    processColumn,
    get_product_type,
    get_process_type,
    get_size,
    preprocess,
)
import glob
import logging
import os
import numpy as np
import psycopg2.extras as extras
from psycopg2 import sql, connect
from io import StringIO
import numpy as np
from pyvi import ViUtils

# FUNCTIONS FOR READING EXCEL FILE
from openpyxl import load_workbook


def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]


def schemaInit(ws):
    schema = pd.DataFrame(iter_rows(ws))
    schema.columns = schema.iloc[0]
    schema = schema.reindex(schema.index.drop(0))
    return schema


def get_schema(fname):
    wb = load_workbook(fname)  # from openpyxl import load_workbook
    ws = wb.active
    schema = schemaInit(ws)
    return schema


# Xử lý file tồn của nhà máy Cà Mau
def preprocessExcel(filename):
    dfsheet1 = pd.read_excel(filename, sheet_name="Sheet1 (2)")
    dfsheet1 = dfsheet1.groupby("Mã vật tư(Mat.Num.)")[
        "Mã vật tư(Mat.Num.)", "Số lượng tồn kho(ĐVT KG2)"
    ].sum()

    df1 = pd.read_excel(filename, sheet_name="Sheet1")

    for i in range(len(df1["Mã vật tư"])):
        for j in range(len(dfsheet1["Mã vật tư(Mat.Num.)"])):
            if str(df1["Mã vật tư"][i]) == str(dfsheet1["Mã vật tư(Mat.Num.)"].iloc[j]):
                df1["Xuất trong kỳ (KG2)"][i] = dfsheet1[
                    "Số lượng tồn kho(ĐVT KG2)"
                ].iloc[j]
                df1["Cuối kỳ (KG2)"][i] = (
                    df1["Đầu kỳ (KG2)"][i]
                    + df1["Nhập trong kỳ (KG2)"][i]
                    - df1["Xuất trong kỳ (KG2)"][i]
                )
    return df1


# Xử lý file excel của danh sách đơn hàng
def orderList(filename):
    df = get_schema(filename)
    df = processColumn(df)
    df = df[df.ma_san_pham_material != ""]
    df = df.replace(r"", np.nan, regex=True)
    return df


# Xử lý file excel của tồn kho
def inventoryMaterial(filename):
    try:
        df = preprocessExcel(filename)
    except:
        df = get_schema(filename)
    df = processColumn(df)
    df = df[df.ngay_in != ""]
    df = df[df.ngay_in != "None"]
    df = df[~df.ma_vat_tu.isnull()]
    df["nha_may"] = df["nha_may"].apply(lambda x: int(x.split("-")[0]))
    df["ma_vat_tu"] = df["ma_vat_tu"].astype(int)
    try:
        df["ngay_in_dau"] = df["ngay_in"].apply(
            lambda x: "-".join((x.split("-")[0]).split("/")[::-1])
        )
        df["ngay_in_cuoi"] = df["ngay_in"].apply(
            lambda x: "-".join((x.split("-")[1]).split("/")[::-1])
        )
    except:
        df["ngay_in_cuoi"] = df["ngay_in"].apply(
            lambda x: "-".join((x.split("/")[::-1]))
        )
        # df["ngay_in_cuoi"] = df["ngay_in"].apply(lambda x: "-".join(x.split("/")[::-1]))
        first_day = "-".join(str(df["ngay_in"].iloc[0]).split("/")[::-1])
        first_day = first_day.split("-")
        first_day[-1] = "1"
        first_day = "-".join(first_day)

        df["ngay_in_dau"] = first_day

    df.drop("ngay_in", axis=1, inplace=True)
    return df


# Xử lý file của thannh pham xuat ban va san xuat
def inoutMaterial(filename):
    df = get_schema(filename)
    df = processColumn(df)
    df = df[df.posting_date != ""]
    df = df.loc[:, ~df.columns.duplicated()]
    df["material"] = df["material"].astype(int)
    try:
        df["quantity_kg2"] = df["quantity_kg2"].astype(float)
        df["quantity"] = df["quantity"].astype(float)
    except:
        df["quantity_kg2"] = df["quantity_kg2"].astype(float)
    return df


# Xử lý file của hợp đồng còn nợ
def inDebtContract(filename):
    df = get_schema(filename)
    renames = dict()
    for col in list(df.columns):
        renames[col] = preprocess(col)
    df.rename(renames, axis=1, inplace=True)
    df = df[df["ten_san_pham_description"] != ""]
    df = df[df["ma_san_pham_material"] != ""]
    df = df.loc[:, ~df.columns.duplicated()]
    df["ngay_tao_bang"] = date.today()
    # df["loai_san_pham"] = df["ten_san_pham_description"].apply(get_product_type)
    # df["loai_che_bien"] = df["ten_san_pham_description"].apply(get_process_type)
    # df["size"] = df["ten_san_pham_description"].apply(get_size)
    return df


# Xử lý file tồn kho chi tiết
def detailedInventory(filename):
    df = get_schema(filename)
    renames = dict()
    for col in list(df.columns):
        renames[col] = preprocess(col)
    df.rename(renames, axis=1, inplace=True)
    df = df[df["ma_vat_tu_mat_num"] != ""]
    df = df[df["ma_vat_tu_mat_num"] != None]
    df = df[~df.ma_vat_tu_mat_num.isnull()]
    df["ngay_tao_bang"] = date.today()

    # df["loai_san_pham"] = df["ten_vat_tu_mat_description"].apply(get_product_type)
    # df["loai_che_bien"] = df["ten_vat_tu_mat_description"].apply(get_process_type)
    # df["size"] = df["ten_vat_tu_mat_description"].apply(get_size)
    return df


# read excel file and dump the rows to db
class DbDump:
    def __init__(
        self,
        username=DATABASE_INFO["user"],
        pasword=DATABASE_INFO["password"],
        host=DATABASE_INFO["host"],
        port=DATABASE_INFO["port"],
        dbname=DATABASE_INFO["database"],
    ):
        self.username = DATABASE_INFO["user"]
        self.password = DATABASE_INFO["password"]
        self.host = DATABASE_INFO["host"]
        self.port = DATABASE_INFO["port"]
        self.dbname = DATABASE_INFO["database"]
        
    #kết nối với cơ sở dữ liệu
    def connect(self, params_dic):
        """ Connect to the PostgreSQL database server """
        conn = None
        try:
            print("Connecting to the PostgreSQL database...")
            conn = psycopg2.connect(**params_dic)
        except (Exception, psycopg2.DatabaseError) as error:
            print(error)
            sys.exit(1)
        print("Connection successful")
        return conn
    
    def singleExcelToDb(self, filename, table):
        '''
        method dump dữ liệu từ file excel vào bảng
        input: file excel, tên bảng
        output: message dữ liệu được đưa vào cơ sở dữ liệu nếu thành công hoặc message thất bại
        '''        
        
        # xóa các bản ghi ngày hôm trước đối với các bảng xuất bán và sản xuất
        deleteStatement = """DELETE FROM {table} WHERE DATE_PART('month',posting_date) = DATE_PART('month',TIMESTAMP '{date}') AND DATE_PART('year',posting_date) = DATE_PART('year',NOW()) AND plant = '{plant}';"""
        # xóa các bản ghi ngày hôm trước đối với các bảng thành phẩm tồn
        deleteStatement1 = """DELETE FROM {table} WHERE DATE_PART('month',ngay_in_cuoi) = DATE_PART('month',TIMESTAMP '{date}') AND DATE_PART('year',ngay_in_cuoi) = DATE_PART('year',NOW()) AND nha_may = '{nha_may}';"""
        # xóa bản ghi cũ đối với danh sách đơn hàng
        deleteStatement2 = """DELETE FROM {table} WHERE nha_may_plant = {nha_may};"""
        # xóa bản ghi cũ đối với các bảng hợp đồng
        deleteStatement3 = """DELETE FROM {table} """

        if table == "thanh_pham_ton_dau_ky":
            df = inventoryMaterial(filename)
            deleteStatement = deleteStatement1.format(
                table=table,
                date=df["ngay_in_cuoi"].iloc[0],
                nha_may=df["nha_may"].iloc[0],
            )

        elif table == "thanh_pham_xuat_ban" or table == "thanh_pham_nhap_san_xuat":
            df = inoutMaterial(filename)
            deleteStatement = deleteStatement.format(
                table=table, date=df["posting_date"].iloc[0], plant=df["plant"].iloc[0],
            )
        elif table == "ds_don_hang_moi":
            df = orderList(filename)
            deleteStatement = deleteStatement3.format(table=table)

        elif table == "hop_dong_con_no":
            df = inDebtContract(filename)
            deleteStatement = deleteStatement3.format(table=table)

        elif table == "thanh_pham_ton_kho_chi_tiet":
            df = detailedInventory(filename)
            deleteStatement = deleteStatement2.format(
                table=table, nha_may=df["nha_may_plant"].iloc[0]
            )
            updateStatement = """UPDATE thanh_pham_ton_kho_chi_tiet SET ton_kho_theo_special_stock_number = NULL WHERE ton_kho_theo_special_stock_number = '';"""

        print("Done reading file, prepare to dump into db ....")
        # chỉ lấy những trường tương ứng với bảng trong csdl
        df = self.getFinalDataFrame(df, table)

        conn = self.connect(DATABASE_INFO)
        self.executeQuery(conn, deleteStatement, "Delete data from yesterday")
        if table == "thanh_pham_ton_kho_chi_tiet":
            self.executeQuery(conn, updateStatement, "Update blank to null")

        self.execute_values(conn, df, table)

    def execute_values(self, conn, df, table):
        """
        Using psycopg2.extras.execute_values() to insert the dataframe
        """
        # Create a list of tupples from the dataframe values
        tuples = [tuple(x) for x in df.to_numpy()]
        # Comma-separated dataframe columns
        cols = ",".join(list(df.columns))
        # SQL quert to execute
        query = "INSERT INTO %s(%s) VALUES %%s" % (table, cols)
        cursor = conn.cursor()
        try:
            extras.execute_values(cursor, query, tuples)
            conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            print("Error: %s" % error)
            conn.rollback()
            cursor.close()
            return 1
        print("execute_values() done")
        cursor.close()
    #method lấy tên các cột trong các bảng
    def get_columns_names(self, table):
        conn = self.connect(DATABASE_INFO)

        # declare an empty list for the column names
        columns = []

        # declare cursor objects from the connection
        col_cursor = conn.cursor()

        # concatenate string for query to get column names
        # SELECT column_name FROM INFORMATION_SCHEMA.COLUMNS WHERE table_name = 'some_table';
        col_names_str = "SELECT column_name FROM INFORMATION_SCHEMA.COLUMNS WHERE "
        col_names_str += "table_name = '{}';".format(table)

        # print the SQL string
        print("\ncol_names_str:", col_names_str)

        try:
            sql_object = sql.SQL(
                # pass SQL statement to sql.SQL() method
                col_names_str
            ).format(
                # pass the identifier to the Identifier() method
                sql.Identifier(table)
            )
            # execute the SQL string to get list with col names in a tuple
            col_cursor.execute(sql_object)
            # get the tuple element from the liast
            col_names = col_cursor.fetchall()
            # print list of tuples with column names
            print("\ncol_names:", col_names)
            # iterate list of tuples and grab first element
            for tup in col_names:
                # append the col name string to the list
                columns += [tup[0]]
            # close the cursor object to prevent memory leaks
            col_cursor.close()

        except Exception as err:
            print("get_columns_names ERROR:", err)
        # return the list of column names
        return columns

    def getFinalDataFrame(self, df, table):
        '''
        method lấy df cuối cùng sau khi đã được tiền xử lý
        input:  df tiền xử lý
        output: df với các cột giống với trong cơ sở dữ liệu
        
        '''
        exist_column = self.get_columns_names(table)
        exist_column = list(set(exist_column))

        try:
            exist_column.remove("id")
        except:
            pass
        df = df[exist_column]
        return df

    def executeQuery(self, conn, query, message):
        try:
            postgresCursor = conn.cursor()
            postgresCursor.execute(query)
            print(query)
            conn.commit()
            print(message)
        except:
            print("Empty query")


def insertMultipleFiles(engine, filepath, table):
    try:
        if len(filepath) == 0:
            print("File not found")
        for attachment in filepath:
            engine.singleExcelToDb(attachment, table)
    except Exception as e:
        print(e)
        print("insert failed")


def main_componentdb_dump():
    today = date.today()
    today = today.strftime("%Y-%m-%d")
    mp_tla = DbDump()

    try:
        xuat_ban = findfiles(
            "BC XUẤT BÁN*.XLSX", f"data/{today}/thanhpham"
        ) + findfiles("03 Báo cáo*.XLSX", f"data/{today}/thanhpham")
        print(xuat_ban)
        insertMultipleFiles(mp_tla, xuat_ban, "thanh_pham_xuat_ban")

    except Exception as e:
        print(e)

    try:
        san_xuat = findfiles(
            "BC THÀNH PHẨM*.XLSX", f"data/{today}/thanhpham"
        ) + findfiles("02 Báo cáo*.XLSX", f"data/{today}/thanhpham")
        print(san_xuat)
        insertMultipleFiles(mp_tla, san_xuat, "thanh_pham_nhap_san_xuat")
    except Exception as e:
        print(e)

    try:
        nhap_xuat_ton = findfiles("NHẬP*.XLSX", f"data/{today}/thanhpham") + findfiles(
            "01 Số liệu*.XLSX", f"data/{today}/thanhpham"
        )
        print(nhap_xuat_ton)
        insertMultipleFiles(mp_tla, nhap_xuat_ton, "thanh_pham_ton_dau_ky")
    except Exception as e:
        print(e)

    try:
        hop_dong_con_no = findfiles("Dư nợ*.XLSX", f"data/{today}/hopdong")
        print(hop_dong_con_no)
        insertMultipleFiles(mp_tla, hop_dong_con_no, "hop_dong_con_no")
    except Exception as e:
        print(e)

    try:
        don_hang_moi = findfiles("Danh sách*.XLSX", f"data/{today}/donhang")
        print(don_hang_moi)
        insertMultipleFiles(mp_tla, don_hang_moi, "ds_don_hang_moi")
    except Exception as e:
        print(e)

    try:
        # ton_kho_chi_tiet = findfiles("Tồn*.XLSX", f"data/{today}/thanhpham")
        ton_kho_chi_tiet = findfiles("Tồn*|.MP*.XLSX", f"data/{today}/thanhpham")

        print(ton_kho_chi_tiet)
        insertMultipleFiles(mp_tla, ton_kho_chi_tiet, "thanh_pham_ton_kho_chi_tiet")
    except Exception as e:
        print(e)


if __name__ == "__main__":
    main_componentdb_dump()
